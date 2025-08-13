import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging
from typing import List, Dict, OrderedDict

# Функция нормализации методов оплаты (удаляет лишние пробелы)
def normalize_method(method: str) -> str:
    return method.strip()

# Основная функция для сохранения данных в Excel
def save_payment_data_to_excel(payment_systems: List[str],
                               withdraw_systems: List[str],
                               payment_names: List[str],
                               withdraw_names: List[str],
                               url: str,
                               login_data: str,
                               currency: str,
                               deposit_count,
                               recommended_methods,
                               ws,
                               start_row: int) -> int:
    # Нормализуем названия методов оплаты и их идентификаторы
    payment_systems = [normalize_method(x) for x in payment_systems]
    withdraw_systems = [normalize_method(x) for x in withdraw_systems]
    payment_names = [normalize_method(x) for x in payment_names]
    withdraw_names = [normalize_method(x) for x in withdraw_names]

    final_methods = []
    all_methods = payment_systems + [method for method in withdraw_systems if method not in payment_systems]
    all_names = payment_names + [name for name in withdraw_names if name not in payment_names]
    print(f"url: {url}")
    
    for i, method in enumerate(all_methods):
        deposit_value = "YES" if method in payment_systems else "NO"
        withdraw_value = "YES" if method in withdraw_systems else "NO"
        associated_name = all_names[i] if i < len(all_names) else ""
        
        status_value = "STAGE" if "stage" in url else "PROD"
        # Добавляем * если метод в топе
        if method in recommended_methods:
            method += "*"

        final_methods.append((method, associated_name, deposit_value, withdraw_value,))

    # Сортируем: сначала рекомендованные методы, затем обычные
    final_methods.sort(key=lambda x: "*" not in x[0])

    logging.debug(f"Объединённый список методов оплаты: {final_methods}")

    # Заголовки таблицы
    headers = ["Paymethod", "Payment Name", "Currency", "Deposit", "Withdraw", "Status"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=start_row + 1, column=col_idx, value=header)

    current_row = start_row + 2
    for method, associated_name, deposit_value, withdraw_value in final_methods:
        ws.cell(row=current_row, column=1, value=method)             # Paymethod
        ws.cell(row=current_row, column=2, value=associated_name)      # Payment Name
        ws.cell(row=current_row, column=3, value=currency)             # Currency
        ws.cell(row=current_row, column=4, value=deposit_value)        # Deposit
        ws.cell(row=current_row, column=5, value=withdraw_value)       # Withdraw
        ws.cell(row=current_row, column=6, value=status_value)
        current_row += 1

    logging.debug(f"Данные успешно записаны до строки {current_row}")
    return current_row

# Функция объединения данных по полю Paymethod

from collections import OrderedDict
from typing import Dict, List, Optional
import openpyxl


def merge_payment_data(
    all_payment_data: Dict[str, Dict[str, Dict[str, str]]],
    login_list: List[str],
    original_order: List[str],
    currency: str,
    recommended_methods: List[str],
    excel_filename: str,
    url: str,
    conditions_map: Optional[Dict[str, str]] = None
) -> OrderedDict:
    merged_data = OrderedDict()
    seen_methods = set()
    ordered_methods = []
    recommended_set = set(recommended_methods)

    # Собираем уникальные title из original_order + actual keys
    for method in original_order:
        if method not in seen_methods:
            ordered_methods.append(method)
            seen_methods.add(method)
    for login in login_list:
        for full_key in all_payment_data.get(login, {}):
            title = full_key.split("|||")[0]
            if title not in seen_methods:
                ordered_methods.append(title)
                seen_methods.add(title)

    for title in ordered_methods:
        deposit = "NO"
        withdraw = "NO"
        names = set()
        details = []
        matched_keys = []

        for login in login_list:
            for full_key, method_data in all_payment_data.get(login, {}).items():
                if not full_key.startswith(f"{title}|||"):
                    continue
                matched_keys.append(full_key)
                if method_data.get("Deposit") == "YES":
                    deposit = "YES"
                if method_data.get("Withdraw") == "YES":
                    withdraw = "YES"
                name = method_data.get("Payment Name", "")
                if name:
                    names.add(name)
                if login not in details:
                    details.append(login)

        full_name = "\n ".join(sorted(names)) if names else ""
        currency_val = currency or "-"
        status = "STAGE" if "stage" in url else "PROD"

        # ⬇️ обновлённая логика Details
        if conditions_map and title in conditions_map:
            detail_str = conditions_map[title]
        else:
            detail_str = "ALL" if len(details) == len(login_list) else "\n ".join(details)

        is_recommended = any(key in recommended_set for key in matched_keys)
        title_out = f"{title}*" if is_recommended else title

        merged_data[title_out] = {
            "Payment Name": full_name,
            "Currency": currency_val,
            "Deposit": deposit,
            "Withdraw": withdraw,
            "Status": status,
            "Details": detail_str
        }

    # Сортируем: рекомендованные сверху
    final = OrderedDict()
    for k in sorted(merged_data, key=lambda x: "*" not in x):
        final[k] = merged_data[k]

    # Пишем в Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged Payments"

    headers = ["Paymethod", "Payment Name", "Currency", "Deposit", "Withdraw", "Status", "Details"]
    ws.append(headers)

    for method, data in final.items():
        row = [method] + list(data.values())
        ws.append(row)

    # Ширина столбцов
    for col, width in {"A": 18, "B": 28, "C": 10, "D": 10, "E": 10, "F": 10, "G": 20}.items():
        ws.column_dimensions[col].width = width

    wb.save(excel_filename)
    return final
