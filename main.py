import argparse
import logging
import openpyxl
import re # Импортируем 're' для работы с регулярными выражениями

# Импортируем вспомогательные функции из поддиректорий
from utils.excel_utils import save_payment_data_to_excel, merge_payment_data
from extractors.base_extractor import BaseExtractor
from extractors.ritzo_extractor import RitzoExtractor
from extractors.rolling_extractor import RollingExtractor
from extractors.nfs_extractor import NeedForSpinExtractor
from extractors.wld_extractor import WildTokyoExtractor
from extractors.godofwins_extractor import GodofwinsExtractor
from extractors.hugo_extractor import HugoExtractor
from extractors.winshark_extractor import WinsharkExtractor
from extractors.spinlander_extractor import SpinlanderExtractor
from extractors.slota_extractor import SlotaExtractor
from extractors.spinline_extractor import SpinlineExtractor
from extractors.glitchspin_extractor import GlitchSpinExtractor
from extractors.azur_extractor import AzurSlotExtractor
from extractors.slotsvader_extractor import SlotsvaderExtractor
from utils.google_drive import create_google_file

# Настройка логирования
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

# Список логинов для тестирования по GEO
geo_groups = {
    "DE": ["0depnoaffdeeurmobi", "0depaffildeeurmobi", "0depaffildeeurdesk", "0depnoaffdeeurdesk", "4depaffildeeurmobi1"],
    "IT": ["0depnoaffiteurmobi", "0depaffiliteurmobi", "0depaffiliteurdesk", "0depnoaffiteurdesk", "4depaffiliteurmobi1"],
    "AT": ["0depnoaffateurmobi", "0depaffilateurmobi", "0depaffilateurdesk", "0depnoaffateurdesk", "4depaffilateurmobi1"],
    "SE": ["0depnoaffseeurmobi", "0depaffilseeurmobi", "0depaffilseeurdesk", "0depnoaffseeurdesk", "4depaffilseeurmobi1"],
    "GR": ["0depnoaffgreurmobi", "0depaffilgreurmobi", "0depaffilgreurdesk", "0depnoaffgreurdesk", "4depaffilgreurmobi1"],
    "IE": ["0depnoaffieeurmobi", "0depaffilieeurmobi", "0depaffilieeurdesk", "0depnoaffieeurdesk", "4depaffilieeurmobi1"],
    "ES": ["0depnoaffeseurmobi", "0depaffileseurmobi", "0depaffileseurdesk", "0depnoaffeseurdesk", "4depaffileseurmobi1"],
    "PT": ["0depnoaffpteurmobi", "0depaffilpteurmobi", "0depaffilpteurdesk1", "0depnoaffpteurdesk", "4depaffilpteurmobi1"],
    "FI": ["0depnoafffieurmobi", "0depaffilfieurmobi", "0depaffilfieurdesk", "0depnoafffieurdesk", "4depaffilfieurmobi1"],
    "DK_DKK": ["0depnoaffdkdkkmobi", "0depaffildkdkkmobi", "0depaffildkdkkdesk", "0depnoaffdkdkkdesk", "4depaffildkdkkmobi1"],
    "DK_EUR": ["0depnoaffdkeurmobi", "0depaffildkeurmobi", "0depaffildkeurdesk", "0depnoaffdkeurdesk", "4depaffildkeurmobi1"],
    "PL_PLN": ["0depnoaffplplnmobi", "0depaffilplplnmobi", "0depaffilplplndesk", "0depnoaffplplndesk", "4depaffilplplnmobi1"],
    "PL_EUR": ["0depnoaffpleurmobi", "0depaffilpleurmobi", "0depaffilpleurdesk", "0depnoaffpleurdesk", "4depaffilpleurmobi1"],
    "CH_CHF": ["0depnoaffchchfmobi", "0depaffilchchfmobi", "0depaffilchchfdesk", "0depnoaffchchfdesk", "4depaffilchchfmobi1"],
    "CH_EUR": ["0depnoaffcheurmobi", "0depaffilcheurmobi", "0depaffilcheurdesk", "0depnoaffcheurdesk", "4depaffilcheurmobi1"],
    "NO_NOK": ["0depnoaffnonokmobi", "0depaffilnonokmobi", "0depaffilnonokdesk", "0depnoaffnonokdesk", "4depaffilnonokmobi1"],
    "NO_EUR": ["0depnoaffnoeurmobi", "0depaffilnoeurmobi", "0depaffilnoeurdesk", "0depnoaffnoeurdesk", "4depaffilnoeurmobi1"],
    "HU_HUF": ["0depnoaffhuhufmobi", "0depaffilhuhufmobi", "0depaffilhuhufdesk", "0depnoaffhuhufdesk", "4depaffilhuhufmobi1"],
    "HU_EUR": ["0depnoaffhueurmobi", "0depaffilhueurmobi", "0depaffilhueurdesk", "0depnoaffhueurdesk", "4depaffilhueurmobi1"],
    "AU_AUD": ["0depnoaffauaudmobi1", "0depaffilauaudmobi1", "0depaffilauauddesk1", "0depnoaffauauddesk1", "4depaffilauaudmobi1"],
    "CA_CAD": ["0depnoaffcacadmobi", "0depaffilcacadmobi", "0depaffilcacaddesk", "0depnoaffcacaddesk", "4depaffilcacadmobi1"]
}

password_data = "123123123"

# Данные для каждого проекта, включая URL для stage и prod
site_list = [
    {
        "name": "Ritzo",
        "stage_url": "https://stage.ritzo.com",
        "prod_url": "https://ritzo.com",
        "extractor_class": RitzoExtractor
    },
    {
        "name": "Rolling",
        "stage_url": "https://stage.rollingslots.com",
        "prod_url": "https://rollingslots.com",
        "extractor_class": RollingExtractor
    },
    {
        "name": "Needforspin",
        "stage_url": "https://stage.needforspin.com",
        "prod_url": "https://needforspin.com",
        "extractor_class": NeedForSpinExtractor
    },
    {
        "name": "Wildtokyo",
        "stage_url": "https://stage.wildtokyo.com",
        "prod_url": "https://wildtokyo.com",
        "extractor_class": WildTokyoExtractor
    },
    {
        "name": "Godofwins",
        "stage_url": "https://stage.godofwins.com",
        "prod_url": "https://godofwins.com",
        "extractor_class": GodofwinsExtractor
    },
    {
        "name": "Hugo",
        "stage_url": "https://stage.hugocasino.com",
        "prod_url": "https://hugocasino.com",
        "extractor_class": HugoExtractor
    },
    {
        "name": "Winshark",
        "stage_url": "https://stage.winshark.com",
        "prod_url": "https://winshark.com",
        "extractor_class": WinsharkExtractor
    },
    {
        "name": "Spinlander",
        "stage_url": "https://stage.spinlander.com",
        "prod_url": "https://spinlander.com",
        "extractor_class": SpinlanderExtractor
    },
    {
        "name": "Slota",
        "stage_url": "https://stage.slota.casino",
        "prod_url": "https://slota.casino",
        "extractor_class": SlotaExtractor
    },
    {
        "name": "Spinline",
        "stage_url": "https://stage.spinline.com",
        "prod_url": "https://spinline.com",
        "extractor_class": SpinlineExtractor
    },
    {
        "name": "Glitchspin",
        "stage_url": "https://stage.glitchspin.com",
        "prod_url": "https://glitchspin.com",
        "extractor_class": GlitchSpinExtractor
    },
    {
        "name": "Azurslot",
        "stage_url": "https://stage.azurslot.com",
        "prod_url": "https://azurslot.com",
        "extractor_class": AzurSlotExtractor
    },
    {
        "name": "Slotsvader",
        "stage_url": "https://stage.slotsvader.com",
        "prod_url": "https://slotsvader.com",
        "extractor_class": SlotsvaderExtractor
    }
]

# Функция для получения базового URL в зависимости от окружения
def get_base_url(project_name, environment):
    """
    Возвращает базовый URL и название окружения для указанного проекта.
    """
    for project in site_list:
        if project["name"] == project_name:
            if environment == "stage":
                 return project["stage_url"], environment
            elif environment == "prod":
                 return project["prod_url"], environment
    raise ValueError(f"Проект {project_name} не найден!")

# Вспомогательные функции для сортировки платежных методов (соответствуют логике фронтенда)
def get_crypto_index(title: str) -> int:
    """
    Возвращает индекс для сортировки криптовалютных методов.
    "Crypto" имеет наивысший приоритет (-1).
    """
    if title == "Crypto":
        return -1
    order = ['USDTT', 'LTC', 'ETH', 'TRX', 'BTC', 'SOL', 'XRP',
             'USDTE', 'DOGE', 'ADA', 'USDC', 'BCH', 'TON']
    for i, prefix in enumerate(order):
        if title.upper().startswith(prefix):
            return i
    return 999 # Для методов, не являющихся криптовалютами из списка

def get_min_dep(details: str) -> int:
    """
    Извлекает минимальное значение DEP (например, "4DEP") из строки деталей.
    Возвращает 99, если DEP не найден.
    """
    matches = re.findall(r'(\d)DEP', details.upper())
    return min(map(int, matches)) if matches else 99

def is_crypto(title: str) -> bool:
    """
    Проверяет, является ли метод криптовалютным, основываясь на его названии.
    """
    return get_crypto_index(title) < 999 or title == "Crypto"

# Основная функция
def main():
    logging.info("Запуск основного скрипта")

    # Создаем парсер для командной строки
    parser = argparse.ArgumentParser(description="Обработка проектов с переключением между окружениями")
    
    # Добавляем параметр командной строки для выбора окружения
    parser.add_argument("--env", choices=["stage", "prod"], default="prod", help="Укажите окружение: stage или prod")
    args = parser.parse_args()

    logging.info(f"Запуск с окружением: {args.env}")

    # Процесс обработки проектов
    for site in site_list:
        site_name = site["name"]
        extractor_class = site["extractor_class"]
        logging.info(f"Обработка проекта: {site_name}")
        
        # Получаем базовый URL и окружение для текущего проекта
        base_url, environment = get_base_url(site_name, args.env)
        logging.info(f"Используем URL: {base_url} и окружение: {environment}")

        # Создаем новую рабочую книгу Excel для каждого проекта
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet']) # Удаляем стандартный лист 'Sheet'
            logging.debug("Стандартный лист 'Sheet' удалён.")

        # Обрабатываем каждый GEO-регион
        for geo_key, login_list in geo_groups.items():
            # Создаем лист для обычных данных авторизации и методов
            ws = wb.create_sheet(title=f"{site_name}_{geo_key}")
            current_start_row = 1 # Начальная строка для записи в текущий лист

            all_payment_data = {} # Инициализируем данные о платежах для текущего GEO
            last_iteration_methods = [] # Методы из последней итерации логина для текущего GEO
            all_recommended_methods = [] # Все рекомендованные методы для текущего GEO

            # Проходим по каждому логину в списке для текущего GEO
            for i, login in enumerate(login_list):
                logging.info(f"Обработка логина: {login} для проекта {site_name}, итерация {i+1}")

                # Определяем User-Agent в зависимости от типа логина (мобильный/десктоп)
                user_agent = ("Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) "
                              "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1"
                              if "mobi" in login else
                              "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.3 Safari/605.1.15")

                # Создаем экземпляр экстрактора для текущего сайта
                extractor = extractor_class(login, password_data, user_agent, base_url)

                if extractor.authenticate():
                    # Получаем платежные и выводные системы
                    (deposit_methods,
                     withdraw_methods,
                     payment_names,
                     withdraw_names,
                     currency,
                     deposit_count,
                     recommended_methods,
                     ) = extractor.get_payment_and_withdraw_systems(geo_key)

                    all_recommended_methods.extend(recommended_methods) # Добавляем рекомендованные методы

                    logging.debug(f"Данные депозита: {deposit_methods}")
                    logging.debug(f"Данные вывода: {withdraw_methods}")

                    # Объединяем методы депозита и вывода, исключая дубликаты
                    combined_methods = deposit_methods + [method for method in withdraw_methods if method not in deposit_methods]
                    
                    # Создаем словарь для сопоставления названий методов и их отображаемых имен
                    method_title_to_payment_name = {}
                    for idx, method_name in enumerate(deposit_methods):
                        if idx < len(payment_names):
                            method_title_to_payment_name[method_name] = payment_names[idx]
                    for idx, method_name in enumerate(withdraw_methods):
                        if idx < len(withdraw_names):
                            # Добавляем только если еще нет, или обновляем если название вывода уникально
                            if method_name not in method_title_to_payment_name:
                                method_title_to_payment_name[method_name] = withdraw_names[idx]

                    # Если это последняя итерация для текущего GEO, сохраняем порядок методов
                    if i == len(login_list) - 1:
                        last_iteration_methods = combined_methods

                    # Записываем данные в текущий лист Excel
                    current_start_row = save_payment_data_to_excel(
                        payment_systems=deposit_methods,
                        withdraw_systems=withdraw_methods,
                        payment_names=payment_names,
                        withdraw_names=withdraw_names,
                        url=base_url,
                        login_data=login,
                        currency=currency,
                        deposit_count=deposit_count,
                        recommended_methods=recommended_methods,
                        ws=ws,
                        start_row=current_start_row
                    )
                    
                    # Сохраняем данные для объединения по всем логинам текущего GEO
                    all_payment_data[login] = {
                        method_title: {
                            "Deposit": "YES" if method_title in deposit_methods else "NO",
                            "Withdraw": "YES" if method_title in withdraw_methods else "NO",
                            "Payment Name": method_title_to_payment_name.get(method_title, ""),
                            "Status": f"{environment.upper()}",
                            "Recommended": "YES" if method_title in recommended_methods else "NO"
                        }
                        for method_title in combined_methods
                    }

                else:
                    logging.error(f"Ошибка авторизации для логина: {login} на проекте {site_name}")

            # Объединяем полученные данные по платежным методам для текущего GEO
            # Примечание: 'currency' будет взята из последнего успешно обработанного логина.
            merged_data = merge_payment_data(
                all_payment_data,
                login_list,
                last_iteration_methods,
                currency, # Валюта из последнего обработанного логина
                all_recommended_methods,
                excel_filename = f"{site_name}_{environment}.xlsx",
                url = base_url
            )
            
            # Создаем новый лист для объединенных и отсортированных данных для данного GEO
            merged_ws = wb.create_sheet(title=f"{site_name}_{geo_key}_merged")

            # Заголовки для объединенного листа
            headers = ["Paymethod", "Payment Name", "Currency", "Deposit", "Withdraw", "Status", "Details"]
            for col_idx, header in enumerate(headers, start=1):
                merged_ws.cell(row=1, column=col_idx, value=header) # Записываем заголовки в первую строку

            # Сортируем методы согласно логике фронтенда
            sorted_methods = sorted(
                merged_data.keys(),
                key=lambda m: (
                    not is_crypto(m),                                             # Не-крипта идет после крипты
                    get_crypto_index(m),                                          # Сортировка внутри криптовалют
                    m not in all_recommended_methods,                             # Рекомендованные методы выше
                    get_min_dep(merged_data[m].get("Details", "")),               # Сортировка по DEP (меньше DEP - выше)
                    merged_data[m].get("Withdraw") == "NO",                       # Методы без вывода в конце
                    m.lower()                                                     # Алфавитная сортировка как запасной вариант
                )
            )

            # Заполняем объединенные данные в отсортированном порядке
            for row_idx, method in enumerate(sorted_methods, start=2): # Начинаем запись данных со второй строки
                data = merged_data[method]
                merged_ws[f"A{row_idx}"] = method
                merged_ws[f"B{row_idx}"] = data.get("Payment Name", "")
                merged_ws[f"C{row_idx}"] = data.get("Currency", "")
                merged_ws[f"D{row_idx}"] = data.get("Deposit", "")
                merged_ws[f"E{row_idx}"] = data.get("Withdraw", "")
                merged_ws[f"F{row_idx}"] = data.get("Status", "")
                merged_ws[f"G{row_idx}"] = data.get("Details", "")
            
        # Сохраняем файл Excel для текущего проекта и окружения
        output_filename = f"{site_name}__{environment}.xlsx"
        wb.save(output_filename)
        logging.info(f"Все данные сохранены в {output_filename}")
        create_google_file(output_filename) # Загружаем файл на Google Drive

if __name__ == "__main__":
    main()
