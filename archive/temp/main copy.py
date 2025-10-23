import openpyxl
import logging
from utils.excel_utils import save_payment_data_to_excel, merge_payment_data, save_merged_data_to_excel
from extractors.ritzo_extractor import RitzoExtractor

# Настройка логирования
logging.basicConfig(
    level=logging.DEBUG,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

de_login_list = [
    "0depnoaffdeeurmobi",
    "0depaffildeeurmobi",
    "0depaffildeeurdesk",
    "0depnoaffdeeurdesk",
    "4depaffildeeurmobi1"
]

geo_groups = {
    "DE": de_login_list
}

site_list = [
    {
        "name": "Ritzo_stage",
        "base_url": "https://stage.ritzo.com/en",
        "is_stage": True
    },
    {
        "name": "Ritzo",
        "base_url": "https://ritzo.com/en",
        "is_stage": False
    }
]

password_data = "123123123"


def main():
    logging.info("Запуск основного скрипта")

    for site in site_list:
        site_name = site["name"]
        base_url = site["base_url"]
        is_stage = site["is_stage"]

        logging.info(f"Обработка сайта: {site_name}")
        wb = openpyxl.Workbook()

        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
            logging.debug("Стандартный лист 'Sheet' удалён.")

        all_payment_data = {}

        for geo_key, login_list in geo_groups.items():
            ws = wb.create_sheet(title=geo_key)
            current_start_row = 1

            for login in login_list:
                logging.info(f"Обработка логина: {login} для сайта {site_name}")

                user_agents = {
                    "mobile_ios": "Mozilla/5.0 (iPhone; CPU iPhone OS 16_6 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.6 Mobile/15E148 Safari/604.1",
                    "mobile_android": "Mozilla/5.0 (Linux; Android 13; SM-G991B) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.5481.153 Mobile Safari/537.36",
                    "desktop_mac": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/14.0.3 Safari/605.1.15",
                    "desktop_windows": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36"
                }

                for agent_type, user_agent in user_agents.items():
                    logging.info(f"Используемый User-Agent ({agent_type}): {user_agent}")

                    extractor = RitzoExtractor(login, password_data, is_stage, user_agent)

                    if extractor.authenticate():
                        deposit_methods, withdraw_methods, payment_names, withdraw_names = extractor.get_payment_and_withdraw_systems()

                        logging.debug(f"Данные депозита: {deposit_methods}")
                        logging.debug(f"Данные вывода: {withdraw_methods}")

                        for method, name in zip(deposit_methods + withdraw_methods, payment_names + withdraw_names):
                            if method not in all_payment_data:
                                all_payment_data[method] = {"Payment Name": name, "Details": set()}
                            all_payment_data[method]["Details"].add(f"{login}_{agent_type}")

                        current_start_row = save_payment_data_to_excel(
                            payment_systems=deposit_methods,
                            withdraw_systems=withdraw_methods,
                            payment_names=payment_names,
                            withdraw_names=withdraw_names,
                            url=base_url,
                            login_data=f"{login}_{agent_type}",
                            ws=ws,
                            start_row=current_start_row
                        )

                    else:
                        logging.error(f"Ошибка авторизации для логина: {login} на сайте {site_name}")

            current_start_row += 3

        output_filename = f"{site_name.lower()}_payment_data.xlsx"
        wb.save(output_filename)
        logging.info(f"Все данные для {site_name} сохранены в {output_filename}")

        ws_merged = wb.create_sheet(title="Merged Data")

        merged_data = {}
        for method, data in all_payment_data.items():
            details = ", ".join(sorted(data["Details"])) if data["Details"] else ""
            merged_data[method] = {
                "Details": details if len(data["Details"]) < len(de_login_list) * 4 else "ALL"
            }

        save_merged_data_to_excel(merged_data, ws_merged, start_row=1)

        merged_output_filename = f"{site_name.lower()}_merged_payment_data.xlsx"
        wb.save(merged_output_filename)
        logging.info(f"Объединённые данные для {site_name} сохранены в {merged_output_filename}")


if __name__ == "__main__":
    main()
